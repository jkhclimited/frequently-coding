import openpyxl as xl
from openpyxl.chart import BarChart, Reference

target_wb = input('Please enter the file name of the workbook you want to open (with extension): ')

def process_workbook(filename):
    wb = xl.load_workbook(filename)
    target_sheet = input('Please enter the name of the target Sheet (Case Sensitive): ')
    sheet = wb[f'{target_sheet}']

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet, 
            min_row=2, 
            max_row=sheet.max_row,
            min_col=4,
            max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb.save(f'{target_wb}.xlsx')

process_workbook(target_wb)